import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests

import openpyxl
import xlsxwriter as xlsxwriter

from PIL import Image
from PyQt5.QtCore import *
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import *
from selenium import webdriver
from bs4 import BeautifulSoup


class FileListThread(QThread):
    sinOut = pyqtSignal(object)

    def __init__(self, key, filename, parent=None):
        super().__init__(parent)
        self.parent = parent
        self.key = key
        self.filename = filename

    def run(self):
        self.file_list(self.key, self.filename)
        self.sinOut.emit(self.key)
        time.sleep(0.001)

    def file_list(self, key, filename):
        for file in os.listdir(filename):
            file_path = os.path.join(filename, file)
            if os.path.isdir(file_path):
                child = QTreeWidgetItem(key)
                child.setText(0, file)
                child.setText(1, file_path)
                child.setIcon(0, QIcon('images/folder.png'))
                self.file_list(child, file_path)
            else:
                child = QTreeWidgetItem(key)
                child.setText(0, file)
                child.setText(1, file_path)
                if file.endswith('.jpg'):
                    child.setIcon(0, QIcon('images/jpg.png'))
                elif file.endswith('.png'):
                    child.setIcon(0, QIcon('images/png.png'))
                elif file.endswith('.zip'):
                    child.setIcon(0, QIcon('images/zip.png'))
                elif file.endswith('.css'):
                    child.setIcon(0, QIcon('images/css.png'))
                elif file.endswith(('.html', '.htm')):
                    child.setIcon(0, QIcon('images/html.png'))
                elif file.endswith('.txt'):
                    child.setIcon(0, QIcon('images/txt.png'))
                else:
                    child.setIcon(0, QIcon('images/file.png'))


class FilePasteThread(QThread):
    '''
    文件粘贴线程
    '''
    sinOut = pyqtSignal(str)

    def __init__(self, source_absurl, des_url, parent=None):
        super(FilePasteThread, self).__init__(parent)
        self.source_absurl = source_absurl
        self.des_url = des_url

    def run(self):
        source_filename = os.path.basename(self.source_absurl)
        des_filename = source_filename
        if source_filename in os.listdir(os.path.dirname(self.des_url)):
            des_filename = '副本-' + source_filename
        des_absurl = os.path.join(os.path.dirname(self.des_url), des_filename)
        with open(self.source_absurl, 'rb') as fs:
            with open(des_absurl, 'wb') as fd:
                fd.write(fs.read())
        self.sinOut.emit(des_filename)


class SavePageUrlThread(QThread):
    sinOut = pyqtSignal(str)

    def __init__(self, url, saveType, fileName, saveLocation, cookies, parent=None):
        super(SavePageUrlThread, self).__init__(parent)
        self.url = url
        self.saveType = saveType
        self.filename = fileName
        self.fileLoacation = saveLocation
        self.cookies = cookies

    def run(self):
        options = webdriver.ChromeOptions()
        # options.add_argument('')
        # print(self.cookies)
        '''打开浏览器实现文件以及内容的读写'''
        self.bro = webdriver.Chrome()
        self.bro.get(self.url)

        '''创建文件名命名的文件夹 以及其内容的分类文件'''
        try:
            file_dir = os.path.join(self.fileLoacation, self.saveType, self.filename)
            os.mkdir(file_dir)
            os.mkdir(os.path.join(file_dir, 'image'))
            os.mkdir(os.path.join(file_dir, 'link'))
            os.mkdir(os.path.join(file_dir, 'text'))
            os.mkdir(os.path.join(file_dir, 'vedio'))
            os.mkdir(os.path.join(file_dir, 'screenshots'))
            self.file_abs_name = os.path.join(file_dir, self.filename + '.html')

        except FileExistsError:
            self.file_abs_name = os.path.join(self.fileLoacation, self.saveType, self.filename, self.filename + '.html')
            pass

        '''页面截屏保存'''
        js_height = self.bro.execute_script("return document.body.clientHeight")  # 页面动态加载后实际高度
        height = self.bro.get_window_rect()['height']  # 浏览器窗口高度
        screen_path = os.path.join(self.fileLoacation, self.saveType, self.filename, 'screenshots')  # 截图保存路径
        screen_dict = {}  # 存放每个屏幕截图的列表，用于合成页面长途
        if js_height % height == 0:
            screen_num = js_height // height
        else:
            screen_num = js_height // height + 1
        for i in range(screen_num):
            self.bro.execute_script(f"window.scrollTo(0, {i * height});")
            time.sleep(0.2)
            filename = os.path.join(screen_path, self.filename + str(i + 1) + '.png')
            self.bro.get_screenshot_as_file(filename)
            screen_dict[str(i)] = Image.open(filename)

        '''网页小姐图合并为长图'''
        img_width,img_height = screen_dict['0'].size
        result = Image.new(screen_dict['0'].mode,(img_width,len(screen_dict)*img_height))
        for i in range(0,len(screen_dict)):
            result.paste(screen_dict[str(i)],box=(0,(i)*img_height))
        result.save(os.path.join(screen_path,self.filename+'.png'))

        page_source = self.bro.page_source
        self.bro.close()


        '''页面解析部分'''
        soup = BeautifulSoup(page_source, 'lxml')

        '''提取链接 进行适当修改补全，实验性功能，不强调100%准确率'''

        link_file_xlsx = os.path.join(self.fileLoacation, self.saveType, self.filename,'link',self.filename+'_link.xlsx') # xlsx文件的绝对路径
        link_file_html = os.path.join(self.fileLoacation, self.saveType, self.filename,'link',self.filename+'_link.html') # html文件的绝对路径
        link_book = xlsxwriter.Workbook(link_file_xlsx) # 创建xlsx文件
        link_book.add_worksheet()  # 为空的xlsx文件添加有一个sheet
        link_book.close()

        link_book = openpyxl.load_workbook(link_file_xlsx)
        link_sheet = link_book[link_book.sheetnames[0]]
        link_sheet.append(['链接href','链接包含文字','链接title属性'])

        link_html_f = open(link_file_html,'wb')
        link_html_f.write(f'<!DOCTYPE html><html><head><title>链接检测html文件</title></head><body><table style="word-break:break-all;word-warp:break-word;width:800px" border=1 cellspacing=0><caption><b>< {self.filename} >链接提取HTML文件</b></caption><tr><th>链接href</th><th>链接内容文字</th><th>链接title</th></tr>'.encode('utf-8'))
        link_list = soup.find_all('a')
        for index,link in enumerate(link_list):
            '''链接href提取并补全'''
            try:
                href = link['href']
                if href.startswith('//'):
                    href = 'https:' + href
                elif href.startswith('/'):
                    schema = self.url.split('//', 1)[0]  # 协议
                    url_root = self.url.split('//')[1].split('/', 1)[0]  # 域名
                    href = '/'.join([schema+'/', url_root]) + href
                    pass
                elif href.startswith('./'):
                    temp_list = self.url.split('/')[:-1]
                    href_split = href.split('./', 1)
                    href = '/'.join(temp_list) + '/' + href_split[1]
                elif href.startswith('../'):  # ../  表示上一层
                    s = '../'
                    href_list = href.split(s)  # 有两个  ../  会被分成3个长度的列表，于是 ../ 的个数为列表长度减一
                    s_num = len(href_list) - 1
                    schema = self.url.split('//', 1)[0]
                    url_list = self.url.split('//', 1)[1].split('/')[:-1]
                    relative_list = url_list[:s_num * -1 - 1]
                    relative_list.insert(0, schema)
                    href = '/'.join(relative_list) + '/' + href_list[-1]
                elif href.startswith('http://') or href.startswith('https://'):
                    pass
                elif 'javascript:' in href:
                    href = None
                elif '(' in href:
                    pass
                else:
                    temp_list = self.url.split('/')[:-1]
                    href = '/'.join(temp_list) + '/' + href
                link['href'] = href
            except KeyError as e:
                href = None

            '''提取连接中的title'''
            try:
                title = link['title']
            except KeyError as e:
                title = '无title属性'

            '''提取链接包裹的文字'''
            try:
                text = link.string
            except Exception as e:
                text = '无文字内容'
            if not text:
                text = '无文字内容'
            if href:
                # print('href ->', href,'text -> ',text,'title -> ',title)
                link_sheet.append([href.strip(),text.strip(),title.strip()])
                tr = f'<tr><th align="left"><a target = "_blank" href="{href.strip()}">{href.strip()}</a></th><th align="left">{text.strip()}</th><th align="left">{title.strip()}</th></tr>\n'
                link_html_f.write(tr.encode('utf-8'))
        link_book.save(link_file_xlsx)
        link_html_f.write('</table></body></html>'.encode('utf-8'))
        link_html_f.close()

        '''提取页面中的img标签并适当补全'''
        imgs = soup.find_all('img')
        images = []  # 存放每个图片的链接的列表
        for img in imgs:
            try:
                src = img['src']
                if src.startswith('//'):
                    src = 'https:' + src
                elif src.startswith('/'):
                    schema = self.url.split('//', 1)[0]  # 协议
                    url_root = self.url.split('//')[1].split('/', 1)[0]  # 域名
                    src = '/'.join([schema + '/', url_root]) + src
                    pass
                elif src.startswith('./'):
                    temp_list = self.url.split('/')[:-1]
                    href_split = src.split('./', 1)
                    src = '/'.join(temp_list) + '/' + href_split[1]
                elif src.startswith('../'):  # ../  表示上一层
                    s = '../'
                    href_list = src.split(s)  # 有两个  ../  会被分成3个长度的列表，于是 ../ 的个数为列表长度减一
                    s_num = len(href_list) - 1
                    schema = self.url.split('//', 1)[0]
                    url_list = self.url.split('//', 1)[1].split('/')[:-1]
                    relative_list = url_list[:s_num * -1 - 1]
                    relative_list.insert(0, schema)
                    src = '/'.join(relative_list) + '/' + href_list[-1]
                elif src.startswith('http://') or src.startswith('https://'):
                    pass
                else:
                    temp_list = self.url.split('/')[:-1]
                    src = '/'.join(temp_list) + '/' + src
                img['src'] = src
                images.append(src)
            except:
                pass
        '''线程池下载页面的图片'''

        threadpool = ThreadPoolExecutor(10)
        # print(images)
        alltask = [threadpool.submit(self.download_img,img) for img in images]
        num = 0  # 记录下载图片的个数，已经名字顺序的索引
        for future in as_completed(alltask):
            res, status_code = future.result()
            try:
                res_url = res.url
            except AttributeError as e:
                continue
            if status_code == 200:
                if res_url.endswith('.png'):
                    img_name = os.path.join(self.fileLoacation, self.saveType, self.filename, 'image',
                                            self.filename + str(num + 1) + '.png')
                    jpg_normal = 1 # 如果 如片故事为正常的jpg 和png 不需转换
                elif res_url.endswith('.webp'):
                    img_name = os.path.join(self.fileLoacation, self.saveType, self.filename, 'image',
                                            self.filename + str(num + 1) + '.jpg')
                    jpg_normal = 0  # 图片结尾为webp的图片需要将webp转换为jpg
                else:
                    img_name = os.path.join(self.fileLoacation, self.saveType, self.filename, 'image',
                                            self.filename + str(num + 1) + '.jpg')
                    jpg_normal = 1
                f = open(img_name,'wb')
                f.write(res.content)
                f.flush()
                f.close()

                def save_image(input_name, output_name):
                    im = Image.open(input_name)
                    if im.mode == "RGBA":
                        im.load()  # required for png.split()
                        background = Image.new("RGB", im.size, (255, 255, 255))
                        background.paste(im, mask=im.split()[3])  # 3 is the alpha channel
                        im = background
                    im.save('{}'.format(output_name), 'JPEG')
                if not jpg_normal:
                    save_image(img_name,img_name)
                num += 1


        '''提取页面css link标签并适当补全'''

        css_links = soup.find_all('link')
        for css_link in css_links:
            try:
                css_href = css_link['href']
                if css_href.startswith('//'):
                    css_href = 'https:' + css_href
                elif css_href.startswith('/'):
                    schema = self.url.split('//', 1)[0]  # 协议
                    url_root = self.url.split('//')[1].split('/', 1)[0]  # 域名
                    css_href = '/'.join([schema + '/', url_root]) + css_href
                    pass
                elif css_href.startswith('./'):
                    temp_list = self.url.split('/')[:-1]
                    href_split = css_href.split('./', 1)
                    css_href = '/'.join(temp_list) + '/' + href_split[1]
                elif css_href.startswith('../'):  # ../  表示上一层
                    s = '../'
                    href_list = css_href.split(s)  # 有两个  ../  会被分成3个长度的列表，于是 ../ 的个数为列表长度减一
                    s_num = len(href_list) - 1
                    schema = self.url.split('//', 1)[0]
                    url_list = self.url.split('//', 1)[1].split('/')[:-1]
                    relative_list = url_list[:s_num * -1 - 1]
                    relative_list.insert(0, schema)
                    css_href = '/'.join(relative_list) + '/' + href_list[-1]
                elif css_href.startswith('http://') or css_href.startswith('https://'):
                    pass
                elif 'javascript:' in css_href:
                    css_href = None
                elif '(' in css_href:
                    pass
                else:
                    temp_list = self.url.split('/')[:-1]
                    css_href = '/'.join(temp_list) + '/' + css_href
                css_link['href'] = css_href
            except KeyError as e:
                pass


        '''提取页面script标签并适当补全'''
        scripts = soup.find_all('script')
        for script in scripts:
            try:
                script_src = script['src']
                if script_src.startswith('//'):
                    script_src = 'https:' + script_src
                elif script_src.startswith('/'):
                    schema = self.url.split('//', 1)[0]  # 协议
                    url_root = self.url.split('//')[1].split('/', 1)[0]  # 域名
                    script_src = '/'.join([schema + '/', url_root]) + script_src
                    pass
                elif script_src.startswith('./'):
                    temp_list = self.url.split('/')[:-1]
                    href_split = script_src.split('./', 1)
                    script_src = '/'.join(temp_list) + '/' + href_split[1]
                elif script_src.startswith('../'):  # ../  表示上一层
                    s = '../'
                    href_list = script_src.split(s)  # 有两个  ../  会被分成3个长度的列表，于是 ../ 的个数为列表长度减一
                    s_num = len(href_list) - 1
                    schema = self.url.split('//', 1)[0]
                    url_list = self.url.split('//', 1)[1].split('/')[:-1]
                    relative_list = url_list[:s_num * -1 - 1]
                    relative_list.insert(0, schema)
                    script_src = '/'.join(relative_list) + '/' + href_list[-1]
                elif script_src.startswith('http://') or script_src.startswith('https://'):
                    script_src = None
                elif 'javascript:' in script_src:
                    script_src = None
                elif '(' in script_src:
                    pass
                else:
                    temp_list = self.url.split('/')[:-1]
                    script_src = '/'.join(temp_list) + '/' + script_src
                script['src']  = script_src
            except KeyError as e:
                pass

        '''提取文字功能'''


        # p_list = soup.find_all('p')
        # for index,p in enumerate(p_list):
        #     print(p.string)

        '''html文件写入操作'''
        html_file = open(self.file_abs_name, 'wb')
        html_file.write(soup.prettify(encoding='utf-8'))
        html_file.flush()
        html_file.close()

        self.sinOut.emit('success')

    def download_img(self,img_url):
        try:
            header = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36'
            }
            res = requests.get(img_url,headers=header ,timeout=10)
            if res.status_code == 200:
                return res,200
            else:
                return '',400
        except Exception as e:
            return ' ',400




